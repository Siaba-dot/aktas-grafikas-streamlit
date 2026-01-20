# app.py – Aktas + Grafikas → X (Pn–Pn) → Periodiškumas → Kaina (TRUNC)
# -*- coding: utf-8 -*-
import io
import re
import calendar
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
# =========================
# THEME / UI
# =========================
st.set_page_config(
    page_title="Aktas + Grafikas → X + Periodiškumas + Kaina (TRUNC)",
    page_icon="🧾",
    layout="wide",
)
NEON_PRIMARY = "#6EE7F9"
NEON_SECOND = "#A78BFA"
BG_GRAD_1 = "#0b0f19"
BG_GRAD_2 = "#12182b"
# CSS – be f-string, naudojam .format(); VISOS CSS { } → dvigubos {{ }}
_css = """
<style>
.stApp {{
  background: linear-gradient(135deg, {BG1} 0%, {BG2} 100%);
  color: #e6eefc; font-family: Inter, Segoe UI, system-ui, -apple-system, sans-serif;
}}
section[data-testid="stSidebar"] {{
  background: #0e1424aa; border-right: 1px solid rgba(255,255,255,0.06);
}}
.stButton>button {{
  background: linear-gradient(90deg, {NP}33, {NS}33);
  color: #eaf7ff; border: 1px solid {NP}55; border-radius: 10px;
  padding: 0.55rem 1rem; transition: all .2s ease;
}}
.stButton>button:hover {{
  box-shadow: 0 0 18px {NP}88, inset 0 0 10px {NS}44; transform: translateY(-1px);
}}
[data-testid="stFileUploaderDropzone"] > div > div {{
  border: 1px dashed {NP}88 !important; background: rgba(255,255,255,0.03);
}}
</style>
""".format(BG1=BG_GRAD_1, BG2=BG_GRAD_2, NP=NEON_PRIMARY, NS=NEON_SECOND)
st.markdown(_css, unsafe_allow_html=True)
st.title("🧾 Aktas + Grafikas → X (Pn–Pn) → Periodiškumas → Kaina (TRUNC)")
st.caption("Formulės lieka. Kaina ir Suma be PVM – be apvalinimo (TRUNC iki 2 skaitmenų). PVM – nereikalingas.")
# =========================
# KONSTANTOS / ŽEMĖLAPIAI
# =========================
HEADER_ROW_INDEX = 8  # Akto antraštės eilutė
WEEKDAYS_LT = ["Pirmadienis", "Antradienis", "Trečiadienis", "Ketvirtadienis", "Penktadienis"]
WD_IDX = {"pirmadienis": 0, "antradienis": 1, "trečiadienis": 2, "treciadienis": 2, "ketvirtadienis": 3, "penktadienis": 4}
SECTION_KEYS = {"i aukštas", "ii aukštas", "iii aukštas", "rūsys", "i aukstas", "ii aukstas", "iii aukstas", "rusys"}
LT_MONTH_GENITIVE = {
 1:"SAUSIO",2:"VASARIO",3:"KOVO",4:"BALANDŽIO",5:"GEGUŽĖS",6:"BIRŽELIO",
 7:"LIEPOS",8:"RUGPJŪČIO",9:"RUGSĖJO",10:"SPALIO",11:"LAPKRIČIO",12:"GRUODŽIO"
}
# Fiksuoti ODS blokai (A1 koordinatės, 5 stulpeliai Pn–Pn)
ODS_GRID_RANGES: Dict[int, str] = {
 11: "C7:G47",  # Lapkritis
 12: "H7:L47", 1: "H7:L47", 2: "H7:L47",  # Gruodis/Sausis/Vasaris
 3: "M7:Q47",  # Kovas
 4: "R7:V47",  # Balandis
 5: "W7:AA47", 6: "W7:AA47", 7: "W7:AA47", 8: "W7:AA47", 9: "W7:AA47",  # Gegužė–Rugsėjis
 10: "AB7:AF47",  # Spalis
}
# .xlsx grafike – mėnesių grupių etiketės
MONTH_GROUP_LABELS = {
 11: ["LAPKRITIS"],
 12: ["GRUODIS- SAUSIS- VASARIS","GRUODIS-SAUSIS-VASARIS","GRUODIS / SAUSIS / VASARIS"],
 1:  ["GRUODIS- SAUSIS- VASARIS","GRUODIS-SAUSIS-VASARIS","GRUODIS / SAUSIS / VASARIS"],
 2:  ["GRUODIS- SAUSIS- VASARIS","GRUODIS-SAUSIS-VASARIS","GRUODIS / SAUSIS / VASARIS"],
 3:  ["KOVAS"],
 4:  ["BALANDIS"],
 5:  ["GEGUŽĖ-BIRŽELIS-LIEPA- RUGPJŪTIS- RUGSĖJIS","GEGUŽĖ–BIRŽELIS–LIEPA–RUGPJŪTIS–RUGSĖJIS","GEGUŽĖ–RUGSĖJIS"],
 6:  ["GEGUŽĖ-BIRŽELIS-LIEPA- RUGPJŪTIS- RUGSĖJIS","GEGUŽĖ–BIRŽELIS–LIEPA–RUGPJŪTIS–RUGSĖJIS","GEGUŽĖ–RUGSĖJIS"],
 7:  ["GEGUŽĖ-BIRŽELIS-LIEPA- RUGPJŪTIS- RUGSĖJIS","GEGUŽĖ–BIRŽELIS–LIEPA–RUGPJŪTIS–RUGSĖJIS","GEGUŽĖ–RUGSĖJIS"],
 8:  ["GEGUŽĖ-BIRŽELIS-LIEPA- RUGPJŪTIS- RUGSĖJIS","GEGUŽĖ–BIRŽELIS–LIEPA–RUGPJŪTIS–RUGSĖJIS","GEGUŽĖ–RUGSĖJIS"],
 9:  ["GEGUŽĖ-BIRŽELIS-LIEPA- RUGPJŪTIS- RUGSĖJIS","GEGUŽĖ–BIRŽELIS–LIEPA–RUGPJŪTIS–RUGSĖJIS","GEGUŽĖ–RUGSĖJIS"],
 10: ["SPALIS"],
}
# Pavadinimų normalizavimo žodynas (grafike -> akte)
# >>> MINIMALUS PATAISYMAS: nepervadinam "Koridoriai, laiptinės, tambūrai, holas"
NAME_OVERRIDE = {}
# Žiemos darbų identifikatoriai (vasaros mėn. blokuoti)
WINTER_KEYWORDS = ["snieg", "salt.laik", "šalt.laik", "salt laik", "šalt laik"]
WINTER_EXACT = {"mechanizuotas sniego stumdymas nuo parkavimo vietų (visame kieme) ir nuo važiuojamosios dalies"}
# =========================
# PAGALBINIAI
# =========================
def norm(s: str) -> str:
    s = (str(s) if s is not None else "").strip().lower()
    s = (s.replace("ą","a").replace("č","c").replace("ę","e").replace("ė","e")
         .replace("į","i").replace("š","s").replace("ų","u").replace("ū","u").replace("ž","z")
         .replace("–","-").replace("—","-"))
    s = re.sub(r"\s+", " ", s)
    return s

def try_parse_date(val) -> Optional[date]:
    if val is None or val == "": return None
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    if isinstance(val, (int, float)):
        try:
            dt = pd.to_datetime(val, unit="D", origin="1899-12-30", errors="coerce")
            if pd.notnull(dt): return dt.date()
        except Exception:
            pass
    if isinstance(val, str):
        s = val.strip()
        for fmt in ("%Y-%m-%d","%m/%d/%Y","%d/%m/%Y","%Y.%m.%d","%d.%m.%Y"):
            try: return datetime.strptime(s, fmt).date()
            except Exception: continue
        dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
        if pd.notnull(dt): return dt.date()
    return None

def build_header_map(ws: Worksheet, header_row: int) -> Dict[str,int]:
    m = {}
    for c in range(1, ws.max_column+1):
        v = ws.cell(header_row, c).value
        if v is None: continue
        m[norm(v)] = c
    return m

def find_end_row(ws: Worksheet, start_row: int) -> int:
    # Stabdome ties "Suma be PVM"
    for r in range(start_row, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and norm(v) == norm("Suma be PVM"):
                return r-1
    return ws.max_row

def detect_name_col(ws: Worksheet, start_row: int, end_row: int, header_map: Dict[str,int]) -> int:
    anchors = [norm("Mato vnt."), norm("įkainis"), norm("Periodiškumas"), norm("ikainis")]
    anchor_cols = [header_map[a] for a in anchors if a in header_map]
    max_anchor = min(anchor_cols) if anchor_cols else ws.max_column
    best_col, best_score = 1, -1
    for c in range(1, max_anchor):
        score = 0
        for r in range(start_row, end_row+1):
            v = ws.cell(r,c).value
            if isinstance(v, str) and v.strip():
                score += 1
        if score > best_score:
            best_col, best_score = c, score
    return best_col

def find_merged_range(ws: Worksheet, row: int, col: int):
    coord = ws.cell(row=row, column=col).coordinate
    for rng in ws.merged_cells.ranges:
        if coord in rng:
            return rng
    return None

def anchor_coords(ws: Worksheet, row: int, col: int) -> Tuple[int,int]:
    rng = find_merged_range(ws, row, col)
    if rng: return rng.min_row, rng.min_col
    return row, col

def write_X_to_weekday(ws: Worksheet, row: int, col: int, value="X"):
    ar, ac = anchor_coords(ws, row, col)
    ws.cell(ar, ac).value = value

def clear_all_weekday_marks(ws: Worksheet, header_map: dict, start_row: int, end_row: int) -> int:
    day_cols = {}
    for label, wd in WD_IDX.items():
        if label in header_map:
            day_cols[wd] = header_map[label]
    cleared = 0
    for r in range(start_row, end_row+1):
        for _, c in day_cols.items():
            ar, ac = anchor_coords(ws, r, c)
            cell = ws.cell(ar, ac)
            if cell.value not in (None, ""):
                cell.value = None
                cleared += 1
    return cleared
# =========================
# ŠVENTĖS
# =========================
def _easter_date(year: int) -> date:
    a = year % 19; b = year // 100; c = year % 100
    d = b // 4; e = b % 4; f = (b + 8) // 25; g = (b - f + 1) // 3
    h = (19*a + b - d - g + 15) % 30; i = c // 4; k = c % 4
    l = (32 + 2*e + 2*i - h - k) % 7; m = (a + 11*h + 22*l) // 451
    month = (h + l - 7*m + 114) // 31; day = ((h + l - 7*m + 114) % 31) + 1
    return date(year, month, day)

def lt_public_holidays(year: int) -> List[date]:
    easter = _easter_date(year)
    easter_mon = pd.Timestamp(easter) + pd.Timedelta(days=1)
    fixed = [
        date(year,1,1), date(year,2,16), date(year,3,11), date(year,5,1),
        date(year,6,24), date(year,7,6), date(year,8,15), date(year,11,1),
        date(year,12,24), date(year,12,25), date(year,12,26),
    ]
    return sorted(set(fixed + [easter, easter_mon.date()]))

def collect_holidays_from_sheet(ws: Worksheet, header_row: int, target_year: Optional[int] = None) -> List[date]:
    if target_year is None:
        y_guess = try_parse_date(ws["A6"].value) if "A6" in ws else None
        if y_guess: target_year = y_guess.year
    target_key = norm("Šventinės dienos")
    anchors = []
    for r in range(1, min(ws.max_row, header_row+120)+1):
        for c in range(1, ws.max_column+1):
            v = ws.cell(r,c).value
            if isinstance(v, str) and norm(v) == target_key:
                anchors.append((r,c))
    bag = []
    for (r0,c0) in anchors:
        for r in range(r0+1, min(ws.max_row, r0+80)+1):
            for c in range(c0, min(ws.max_column, c0+8)+1):
                v = ws.cell(r,c).value
                if v is None: continue
                d = try_parse_date(v)
                if d: bag.append(d); continue
                if isinstance(v, str):
                    s = v.replace(",", " ").replace(";", " ")
                    for tok in s.split():
                        d = try_parse_date(tok)
                        if d: bag.append(d); continue
                        m = re.match(r"^(\d{1,2})\.(\d{1,2})$", tok.strip())
                        if m and target_year:
                            day = int(m.group(1)); mon = int(m.group(2))
                            try: bag.append(date(target_year, mon, day))
                            except Exception: pass
    return sorted(set(bag))
# =========================
# GRAFIKO NUSKAITYMAS
# =========================
def col_letter_to_index(letters: str) -> int:
    letters = letters.strip().upper(); idx = 0
    for ch in letters:
        if not ch.isalpha(): break
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1

def parse_a1_range(a1: str) -> Tuple[int,int,int,int]:
    a1 = a1.replace(" ",""); left,right = a1.split(":")
    m1 = re.match(r"([A-Za-z]+)(\d+)", left); m2 = re.match(r"([A-Za-z]+)(\d+)", right)
    c1,r1 = m1.group(1), int(m1.group(2)); c2,r2 = m2.group(1), int(m2.group(2))
    col_start = col_letter_to_index(c1); col_end = col_letter_to_index(c2)
    row_start = r1 - 1; row_end = r2 - 1
    return row_start, row_end, col_start, col_end

def read_schedule_ods(ods_file, month: int) -> List[Dict]:
    try:
        df = pd.read_excel(ods_file, engine="odf", header=None)
    except Exception:
        st.error("Nepavyko atidaryti ODS (`odfpy` paketas) arba failas sugadintas.")
        raise
    a1 = ODS_GRID_RANGES.get(month)
    if not a1:
        raise RuntimeError(f"Mėnesiui {month} ODS blokas nesukonfigūruotas.")
    r0, r1, c0, c1 = parse_a1_range(a1)
    block_cols = list(range(c0, c0 + 5))  # 5 stulpeliai: Pn–Pn
    # Pavadinimų stulpelis kairėje nuo bloko
    name_col = None
    best_score = -1
    for cc in range(c0 - 1, -1, -1):
        score = 0
        for rr in range(r0, r1 + 1):
            v = df.iat[rr, cc] if cc < df.shape[1] else None
            if isinstance(v, str) and v.strip():
                nv = norm(v)
                if nv not in {"pirmadienis","antradienis","treciadienis","ketvirtadienis","penktadienis"} and not nv.isdigit():
                    score += 1
        if score > best_score:
            best_score = score
            name_col = cc
    if name_col is None:
        name_col = 0
    items: List[Dict] = []
    current_section: Optional[str] = None
    for rr in range(r0, r1 + 1):
        raw = df.iat[rr, name_col] if name_col < df.shape[1] else None
        if not isinstance(raw, str) or not raw.strip():
            continue
        name = raw.strip()
        nn = norm(name)
        if nn in SECTION_KEYS:
            current_section = name
            continue
        wdset: set = set()
        override: Optional[int] = None
        for i, cc in enumerate(block_cols):
            if cc >= df.shape[1]:
                continue
            v = df.iat[rr, cc]
            s = str(v).strip().lower() if v is not None else ""
            if s == "x" or s.startswith("x") or "×" in s:
                wdset.add(i)
            m = re.search(r"(\d+)\s*kart", s)
            if m:
                override = int(m.group(1))
        # paieška aplink pavadinimą
        for cc in range(max(0, name_col - 3), min(df.shape[1] - 1, name_col + 6) + 1):
            v = df.iat[rr, cc]
            if isinstance(v, str):
                m = re.search(r"(\d+)\s*kart", v.lower())
                if m:
                    override = int(m.group(1))
        items.append({
            "name": name,
            "section": current_section,  # I aukštas / II aukštas / III aukštas / Rūsys
            "wdset": (wdset if wdset else None),  # X savaitės dienos (0..4)
            "override": override  # N kartai per mėn., jei nurodyta
        })
    return items

def read_schedule_xlsx(xlsx_file, month: int) -> List[Dict]:
    wb = load_workbook(xlsx_file, data_only=True)
    ws = wb.active
    labels = MONTH_GROUP_LABELS.get(month, [])
    anchor = None
    for r in range(1, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            v = ws.cell(r,c).value
            if not isinstance(v, str): continue
            nv = norm(v)
            if any(norm(lbl) == nv for lbl in labels):
                anchor = (r,c); break
        if anchor: break
    if not anchor:
        raise RuntimeError("Nerasta mėnesio grupės antraštė grafike (.xlsx).")
    WEEK = [norm(x) for x in WEEKDAYS_LT]
    block_row=None; block_cols=None
    for rr in range(anchor[0], min(ws.max_row, anchor[0]+8)+1):
        for cc in range(anchor[1], min(ws.max_column-4, anchor[1]+25)+1):
            lab = [ws.cell(rr, cc+i).value for i in range(5)]
            if all(isinstance(x, str) for x in lab) and [norm(x) for x in lab] == WEEK:
                block_row=rr; block_cols=[cc+i for i in range(5)]; break
        if block_row: break
    if not block_row:
        raise RuntimeError("Pn–Pn antraštės nerastos nurodytame bloke (.xlsx).")
    # Pavadinimų stulpelis – kairėje nuo bloko, labiausiai „tekstinis“
    name_col=None; best=-1
    for cc in range(block_cols[0]-1, 0, -1):
        score = sum(1 for r in range(block_row+1, min(ws.max_row, block_row+200)+1)
                    if isinstance(ws.cell(r,cc).value, str) and ws.cell(r,cc).value.strip())
        if score>best: best=score; name_col=cc
    items=[]; current_section=None
    for r in range(block_row+1, ws.max_row+1):
        raw = ws.cell(r, name_col).value
        if not isinstance(raw, str) or not raw.strip(): continue
        name = raw.strip()
        nn = norm(name)
        if nn in SECTION_KEYS:
            current_section = name
            continue
        wdset=set(); override=None
        for i,c in enumerate(block_cols):
            v = ws.cell(r,c).value
            s = str(v).strip().lower() if v is not None else ''
            if s=='x' or s.startswith('x') or '×' in s: wdset.add(i)
            m = re.search(r"(\d+)\s*kart", s)
            if m: override=int(m.group(1))
        # „N kartai…“ šalia pavadinimo
        for cc in range(max(1, name_col-3), min(ws.max_column, name_col+10)+1):
            v = ws.cell(r,cc).value
            if isinstance(v, str):
                m = re.search(r"(\d+)\s*kart", v.lower())
                if m: override=int(m.group(1))
        items.append({"name": name, "section": current_section, "wdset": (wdset if wdset else None), "override": override})
    return items
# =========================
# AKTO EILUTĖS / MAP
# =========================
def extract_act_rows(ws: Worksheet, header_map: Dict[str,int]):
    start = HEADER_ROW_INDEX + 1
    end = find_end_row(ws, start)
    name_col = detect_name_col(ws, start, end, header_map)
    rows=[]; current_section=None
    for r in range(start, end+1):
        v = ws.cell(r, name_col).value
        if isinstance(v, str) and v.strip():
            nm = v.strip()
            nn = norm(nm)
            if nn in SECTION_KEYS:
                current_section = nm
                continue
            rows.append({"row": r, "name": nm, "section": current_section})
    day_cols={}
    for label, wd in WD_IDX.items():
        if label in header_map: day_cols[wd] = header_map[label]
    return start, end, name_col, rows, day_cols

def is_winter_only_task(name: str) -> bool:
    n = norm(name)
    if n in WINTER_EXACT: return True
    return any(k in n for k in WINTER_KEYWORDS)

def token_similarity(a: str, b: str) -> float:
    ta = set(norm(a).split()); tb = set(norm(b).split())
    if not ta or not tb: return 0.0
    return len(ta & tb) / len(ta | tb)

def map_schedule_to_act(sched_items: List[Dict], act_rows: List[Dict], name_override: Dict[str,str], min_score: float = 0.6) -> List[Dict]:
    def name_norm_fix(s: str) -> str:
        # pašalinam klaidinančius tarpus prieš ')'
        return re.sub(r"\)\s+$", ")", s)
    pairs=[]
    for it in sched_items:
        sname = name_norm_fix(name_override.get(it["name"], it["name"]))
        # kandidatai – tik tame pačiame aukšte
        cands = [ar for ar in act_rows if ar['section'] == it['section']]
        # tiksli lygybė
        exact=None
        for ar in cands:
            if norm(ar['name']) == norm(sname): exact = ar; break
        if exact is not None:
            pairs.append({**it, "act_row": exact['row'], "act_name": exact['name'], "act_section": exact['section'], "score": 1.0})
            continue
        # fuzzy tame aukšte
        best=None; best_s=-1.0
        for ar in cands:
            s = token_similarity(sname, ar['name'])
            if s > best_s: best, best_s = ar, s
        # global fallback tik jei tame aukšte nerasta
        if (best is None or best_s < min_score) and not cands:
            for ar in act_rows:
                s = token_similarity(sname, ar['name'])
                if s > best_s: best, best_s = ar, s
        if best and best_s >= min_score:
            pairs.append({**it, "act_row": best['row'], "act_name": best['name'], "act_section": best['section'], "score": best_s})
    return pairs
# =========================
# TAIKYMAS AKTUI
# =========================
def compute_period_for_row(ws: Worksheet, r: int, day_cols: Dict[int,int], year: int, month: int, holidays: List[date]) -> int:
    wdset=set()
    for wd, c in day_cols.items():
        ar, ac = anchor_coords(ws, r, c)
        v = ws.cell(ar, ac).value
        if v is None: continue
        s = str(v).strip().lower()
        if s=='x' or s.startswith('x') or '×' in s:
            wdset.add(wd)
    if not wdset: return 0
    holi = {d for d in holidays if d.year == year and d.month == month}
    last_day = calendar.monthrange(year, month)[1]
    cnt=0
    for d in range(1, last_day+1):
        dt = date(year, month, d)
        if dt.weekday() in wdset and dt not in holi:
            cnt+=1
    return cnt

def apply_schedule_to_act(
    wb: Workbook,
    year: int,
    month: int,
    sched_items: List[Dict],
    skip_winter_in_summer: bool = True,
    union_public_holidays: bool = True,
    exclude_holidays: bool = True,
    min_match_score: float = 0.6
):
    ws = wb.active
    header_map = build_header_map(ws, HEADER_ROW_INDEX)
    start, end, name_col, act_rows, day_cols = extract_act_rows(ws, header_map)
    period_col = header_map.get(norm("Periodiškumas"))
    plotas_col = header_map.get(norm("Plotas kv m./kiekis/val"))
    ikainis_col = header_map.get(norm("įkainis")) or header_map.get(norm("ikainis"))
    kaina_col = header_map.get(norm("Kaina"))
    pairs = map_schedule_to_act(sched_items, act_rows, NAME_OVERRIDE, min_match_score)
    # prieš rašant – išvalom visas Pn–Pn žymas
    clear_all_weekday_marks(ws, header_map, start, end)
    # švenčių sąrašas
    holidays: List[date] = []
    if exclude_holidays:
        holidays = collect_holidays_from_sheet(ws, HEADER_ROW_INDEX, target_year=year)
    if union_public_holidays:
        holidays = sorted(set(holidays + lt_public_holidays(year)))
    # rašom overrides arba X
    x_written = 0; overrides_set = 0
    for p in pairs:
        r = p['act_row']
        nm = p['name']
        # vasarą – žiemos darbų nerašom
        if skip_winter_in_summer and month in (5,6,7,8,9) and is_winter_only_task(nm):
            if period_col: ws.cell(r, period_col).value = 0
            continue
        if p['override'] is not None:
            if period_col: ws.cell(r, period_col).value = int(p['override']); overrides_set += 1
            continue
        if p['wdset']:
            for wd in p['wdset']:
                c = day_cols.get(wd)
                if c:
                    write_X_to_weekday(ws, r, c, "X")
                    x_written += 1
    # periodiškumai (jei override > 0 – paliekam)
    updated = 0
    last_day = calendar.monthrange(year, month)[1]
    holi = {d for d in holidays if d.year==year and d.month==month}
    wd_counts = {i:0 for i in range(5)}
    for d in range(1, last_day+1):
        dt = date(year, month, d); wd = dt.weekday()
        if wd < 5 and dt not in holi: wd_counts[wd]+=1
    if period_col:
        for r in range(start, end+1):
            v = ws.cell(r, period_col).value
            if isinstance(v, (int,float)) and int(v) > 0:
                updated += 1
                continue
            newp = compute_period_for_row(ws, r, day_cols, year, month, holidays)
            ws.cell(r, period_col).value = int(newp)
            updated += 1
    # kainų formulės
    if kaina_col and plotas_col and ikainis_col and period_col:
        for r in range(start, end+1):
            pl = ws.cell(r, plotas_col).coordinate
            ik = ws.cell(r, ikainis_col).coordinate
            pe = ws.cell(r, period_col).coordinate
            cell = ws.cell(r, kaina_col)
            cell.value = f"=TRUNC({pl}*{ik}*{pe},2)"
            cell.number_format = "0.00"
    # Suma be PVM
    sum_row = None
    for r in range(start, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            v = ws.cell(r,c).value
            if isinstance(v, str) and norm(v) == norm("Suma be PVM"):
                sum_row = r; break
        if sum_row: break
    if sum_row and kaina_col:
        rng = f"{ws.cell(start, kaina_col).coordinate}:{ws.cell(end, kaina_col).coordinate}"
        cell = ws.cell(sum_row, kaina_col)
        cell.value = f"=TRUNC(SUM({rng}),2)"
        cell.number_format = "0.00"
    return updated, x_written, overrides_set, wd_counts
# =========================
# UI
# =========================
with st.sidebar:
    st.header("⚙️ Nustatymai")
    c1, c2 = st.columns(2)
    with c1:
        target_year = st.number_input("Metai", 2020, 2100, datetime.now().year, step=1)
    with c2:
        target_month = st.number_input("Mėnuo", 1, 12, datetime.now().month, step=1)
    date_format = st.selectbox("Datos formatas A6 langelyje", ["MM/DD/YYYY","YYYY-MM-DD","DD.MM.YYYY","YYYY.MM.DD"], index=0)
    exclude_holidays = st.checkbox("Neįtraukti švenčių (iš skyriaus „Šventinės dienos“)", value=True)
    union_public_holidays = st.checkbox("Papildomai naudoti LT nedarbo dienas (automatiškai)", value=True)
    skip_winter_in_summer = st.checkbox("Vasarą nerodyti žieminių darbų (sniegas, šalt.laik.)", value=True)
    clear_existing_x = st.checkbox("Išvalyti SENAS Pn–Pn žymas", value=True)
    min_match_score = st.slider("Mažiausias 'fuzzy' atitikimo balas", 0.40, 0.95, 0.60, 0.05)

st.subheader("1) Įkelk aktą (.xlsx) ir grafiką (.ods ARBA .xlsx)")
act_file = st.file_uploader("Aktas (Excel .xlsx)", type=["xlsx"])
sched_file = st.file_uploader("Grafikas (.ods arba .xlsx)", type=["ods","xlsx"])

if st.button("🔄 Grafikas → X (Pn–Pn) → Periodiškumas → Kaina", type="primary", use_container_width=True):
    if not act_file:
        st.warning("Įkelk aktą (.xlsx).")
        st.stop()
    if not sched_file:
        st.warning("Įkelk grafiką (.ods arba .xlsx).")
        st.stop()
    try:
        # atveriam aktą
        wb = load_workbook(filename=act_file, data_only=False)
        ws = wb.active
        # A6/C7 antraštės datos
        last_day = calendar.monthrange(int(target_year), int(target_month))[1]
        fmt_map = {"MM/DD/YYYY":"%m/%d/%Y","YYYY-MM-DD":"%Y-%m-%d","DD.MM.YYYY":"%d.%m.%Y","YYYY.MM.DD":"%Y.%m.%d"}
        ws["A6"].value = date(int(target_year), int(target_month), last_day).strftime(fmt_map.get(date_format, "%m/%d/%Y"))
        ws["C7"].value = f"{LT_MONTH_GENITIVE[int(target_month)]} 1-{last_day}"
        # jei pažymėta – išvalom visas senas X
        header_map = build_header_map(ws, HEADER_ROW_INDEX)
        start = HEADER_ROW_INDEX + 1
        end = find_end_row(ws, start)
        if clear_existing_x:
            _ = clear_all_weekday_marks(ws, header_map, start, end)
        # nuskaityti grafiką
        if sched_file.name.lower().endswith(".ods"):
            sched_items = read_schedule_ods(sched_file, int(target_month))
        else:
            sched_items = read_schedule_xlsx(sched_file, int(target_month))
        # pritaikyti
        updated, x_written, overrides_set, wd_counts = apply_schedule_to_act(
            wb=wb, year=int(target_year), month=int(target_month),
            sched_items=sched_items,
            skip_winter_in_summer=skip_winter_in_summer,
            union_public_holidays=union_public_holidays,
            exclude_holidays=exclude_holidays,
            min_match_score=float(min_match_score),
        )
        # išsaugoti išvestį
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        label = "{:04d}-{:02d}".format(int(target_year), int(target_month))
        st.success(
            "✔ Nauji X parašyti: {x}\n"
            "✔ Periodiškumas atnaujintas: {u}\n"
            "✔ Periodinių (N kart./mėn.) įrašyta: {o}\n"
            "Pn: {m0}, An: {m1}, Tr: {m2}, Kt: {m3}, Pnkt: {m4}".format(
                x=x_written, u=updated, o=overrides_set,
                m0=wd_counts.get(0,0), m1=wd_counts.get(1,0),
                m2=wd_counts.get(2,0), m3=wd_counts.get(3,0), m4=wd_counts.get(4,0)
            )
        )
        st.download_button(
            "⬇️ Parsisiųsti atnaujintą aktą",
            data=out,
            file_name=f"Aktas_atnaujintas_{label}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    except Exception as e:
        st.exception(e)
        st.error("Nepavyko pažymėti X / perskaičiuoti / įrašyti formulių. Patikrink grafiko mėnesio antraštes ir akto stulpelių pavadinimus.")
